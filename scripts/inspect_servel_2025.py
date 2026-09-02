from __future__ import annotations

import io
import json
import zipfile
from pathlib import Path

import requests
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "data" / "legislative" / "2026" / "affiliations"
URL = "https://www.servel.cl/wp-content/uploads/2025/11/PRELIMINARES_DIPUTADOS.zip"


def cell_value(value):
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        return value
    return str(value).strip()


def inspect_xlsx(raw: bytes) -> dict:
    workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    sheets = []
    for sheet in workbook.worksheets[:3]:
        rows = []
        for row in sheet.iter_rows(min_row=1, max_row=35, max_col=16, values_only=True):
            values = [cell_value(x) for x in row]
            if any(x != "" for x in values):
                rows.append(values)
        sheets.append({"title": sheet.title, "preview_rows": rows})
    return {"sheet_names": workbook.sheetnames, "sheets": sheets}


def main() -> None:
    OUT.mkdir(parents=True, exist_ok=True)
    session = requests.Session()
    session.headers.update({"User-Agent": "Mozilla/5.0 conoce-a-tu-parlamentario/2.0"})
    response = session.get(URL, timeout=120)
    response.raise_for_status()
    archive = zipfile.ZipFile(io.BytesIO(response.content))

    names = [x for x in archive.namelist() if x.lower().endswith(".xlsx")]
    # Para conocer el contrato basta revisar distritos separados geográficamente.
    sample_names = []
    for target in ("DISTRITO_1.xlsx", "DISTRITO_8.xlsx", "DISTRITO_20.xlsx", "DISTRITO_28.xlsx"):
        hit = next((x for x in names if x.upper().endswith(target)), None)
        if hit:
            sample_names.append(hit)

    samples = []
    for name in sample_names:
        raw = archive.read(name)
        samples.append({
            "name": name,
            "size": len(raw),
            **inspect_xlsx(raw),
        })

    diagnostics = {
        "source_url": URL,
        "download_bytes": len(response.content),
        "xlsx_files": len(names),
        "sample_files": samples,
    }
    (OUT / "servel_2025_zip_diagnostics.json").write_text(
        json.dumps(diagnostics, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    print(json.dumps(diagnostics, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
