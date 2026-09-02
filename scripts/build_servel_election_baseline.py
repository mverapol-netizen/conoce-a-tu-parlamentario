from __future__ import annotations

import csv
import io
import json
import re
import unicodedata
import zipfile
from collections import defaultdict
from difflib import SequenceMatcher
from pathlib import Path

import requests
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
PROFILES = ROOT / "assets" / "js" / "profiles.js"
OUT = ROOT / "data" / "legislative" / "2026" / "affiliations"
BASELINE = OUT / "election_baseline_2025.csv"
DIAGNOSTICS = OUT / "election_baseline_diagnostics.json"
SOURCE_URL = "https://www.servel.cl/wp-content/uploads/2025/11/PRELIMINARES_DIPUTADOS.zip"

FIELDS = [
    "deputy_id", "deputy_name", "district", "servel_candidate_code", "servel_candidate_name",
    "electoral_party", "electoral_pact", "electoral_subpact", "source_url",
    "match_method", "match_score", "match_confidence",
]


def norm(value: str) -> str:
    value = unicodedata.normalize("NFD", str(value or ""))
    value = "".join(ch for ch in value if unicodedata.category(ch) != "Mn")
    value = value.lower().replace("ñ", "n")
    value = re.sub(r"[^a-z0-9 ]+", " ", value)
    return re.sub(r"\s+", " ", value).strip()


def name_tokens(value: str) -> set[str]:
    return {x for x in norm(value).split() if len(x) > 1}


def load_profiles() -> list[dict]:
    text = PROFILES.read_text(encoding="utf-8")
    match = re.search(r"window\.PROFILES\s*=\s*(\{.*\})\s*;\s*$", text, re.S)
    if not match:
        raise RuntimeError("No se pudo leer window.PROFILES")
    raw = json.loads(match.group(1))
    rows = []
    for key, profile in raw.items():
        rows.append({
            "deputy_id": str(profile.get("id") or ""),
            "deputy_name": profile.get("officialName") or key,
            "district": int(profile.get("district") or 0),
        })
    if len(rows) != 155:
        raise RuntimeError(f"Se esperaban 155 perfiles; hay {len(rows)}")
    return rows


def download_archive() -> zipfile.ZipFile:
    session = requests.Session()
    session.headers.update({"User-Agent": "Mozilla/5.0 conoce-a-tu-parlamentario/2.0"})
    response = session.get(SOURCE_URL, timeout=180)
    response.raise_for_status()
    return zipfile.ZipFile(io.BytesIO(response.content))


def read_elected(archive: zipfile.ZipFile) -> list[dict]:
    candidates: dict[str, dict] = {}
    xlsx_names = sorted(x for x in archive.namelist() if x.lower().endswith(".xlsx"))
    if len(xlsx_names) != 28:
        raise RuntimeError(f"Se esperaban 28 archivos de distrito; hay {len(xlsx_names)}")

    for filename in xlsx_names:
        raw = archive.read(filename)
        workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        sheet = workbook.active
        iterator = sheet.iter_rows(values_only=True)
        header = next(iterator, None)
        if not header:
            continue
        columns = {norm(str(value or "")): idx for idx, value in enumerate(header)}
        required = {
            "distrito", "partido", "pacto", "subpacto", "cod candidato",
            "nombre candidato", "electo nominado",
        }
        # Los encabezados originales usan guion bajo; norm() lo convierte en espacios.
        missing = required - set(columns)
        if missing:
            raise RuntimeError(f"{filename}: faltan columnas {sorted(missing)}")

        for row in iterator:
            def get(field: str):
                idx = columns[field]
                return row[idx] if idx < len(row) else None

            elected_raw = get("electo nominado")
            try:
                elected = int(elected_raw or 0) == 1
            except (TypeError, ValueError):
                elected = str(elected_raw or "").strip() == "1"
            if not elected:
                continue

            code = str(get("cod candidato") or "").strip()
            if not code:
                continue
            district_text = str(get("distrito") or "")
            district_match = re.search(r"(\d+)", district_text)
            district = int(district_match.group(1)) if district_match else 0
            record = {
                "district": district,
                "servel_candidate_code": code,
                "servel_candidate_name": str(get("nombre candidato") or "").strip(),
                "electoral_party": str(get("partido") or "").strip(),
                "electoral_pact": str(get("pacto") or "").strip(),
                "electoral_subpact": str(get("subpacto") or "").strip(),
            }
            previous = candidates.get(code)
            if previous and previous != record:
                raise RuntimeError(f"Datos inconsistentes para candidato {code}: {previous} vs {record}")
            candidates[code] = record

    result = sorted(candidates.values(), key=lambda x: (x["district"], x["servel_candidate_name"]))
    if len(result) != 155:
        raise RuntimeError(f"Servel no produjo 155 electos únicos; produjo {len(result)}")
    return result


def similarity(profile: dict, candidate: dict) -> float:
    a = norm(profile["deputy_name"])
    b = norm(candidate["servel_candidate_name"])
    seq = SequenceMatcher(None, a, b).ratio()
    ta, tb = name_tokens(a), name_tokens(b)
    token = len(ta & tb) / max(len(ta | tb), 1)
    containment = len(ta & tb) / max(min(len(ta), len(tb)), 1)
    return round(0.45 * seq + 0.25 * token + 0.30 * containment, 4)


def match_profiles(profiles: list[dict], elected: list[dict]) -> tuple[list[dict], list[dict], list[dict]]:
    elected_by_district: dict[int, list[dict]] = defaultdict(list)
    for candidate in elected:
        elected_by_district[candidate["district"]].append(candidate)

    used_codes: set[str] = set()
    matches = []
    ambiguous = []
    unmatched = []

    for profile in sorted(profiles, key=lambda x: (x["district"], x["deputy_name"])):
        pool = [x for x in elected_by_district[profile["district"]] if x["servel_candidate_code"] not in used_codes]
        exact = [x for x in pool if norm(x["servel_candidate_name"]) == norm(profile["deputy_name"])]

        chosen = None
        method = ""
        score = 0.0
        confidence = ""

        if len(exact) == 1:
            chosen = exact[0]
            method = "district_exact_normalized_name"
            score = 1.0
            confidence = "high"
        else:
            ranked = sorted(((similarity(profile, x), x) for x in pool), key=lambda pair: pair[0], reverse=True)
            if ranked:
                best_score, best = ranked[0]
                second_score = ranked[1][0] if len(ranked) > 1 else 0.0
                margin = best_score - second_score
                # Conservador: nombres deben ser muy parecidos y separarse claramente del segundo candidato del distrito.
                if best_score >= 0.78 and margin >= 0.08:
                    chosen = best
                    method = "district_fuzzy_name"
                    score = best_score
                    confidence = "high" if best_score >= 0.90 else "medium"
                else:
                    ambiguous.append({
                        "profile": profile,
                        "best": best,
                        "best_score": best_score,
                        "second_score": second_score,
                    })

        if chosen is None:
            unmatched.append(profile)
            continue

        used_codes.add(chosen["servel_candidate_code"])
        matches.append({
            "deputy_id": profile["deputy_id"],
            "deputy_name": profile["deputy_name"],
            "district": profile["district"],
            **chosen,
            "source_url": SOURCE_URL,
            "match_method": method,
            "match_score": f"{score:.4f}",
            "match_confidence": confidence,
        })

    unused = [x for x in elected if x["servel_candidate_code"] not in used_codes]
    matches.sort(key=lambda x: int(x["deputy_id"]))
    return matches, ambiguous, unused


def main() -> None:
    OUT.mkdir(parents=True, exist_ok=True)
    profiles = load_profiles()
    elected = read_elected(download_archive())
    matches, ambiguous, unused = match_profiles(profiles, elected)

    with BASELINE.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=FIELDS, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(matches)

    diagnostics = {
        "source_url": SOURCE_URL,
        "servel_elected_unique": len(elected),
        "current_profiles": len(profiles),
        "matched": len(matches),
        "high_confidence_matches": sum(x["match_confidence"] == "high" for x in matches),
        "medium_confidence_matches": sum(x["match_confidence"] == "medium" for x in matches),
        "ambiguous": ambiguous,
        "unmatched_profiles": [x for x in profiles if x["deputy_id"] not in {m["deputy_id"] for m in matches}],
        "unused_servel_elected": unused,
        "electoral_party_counts": dict(sorted(defaultdict(int, {k: sum(1 for x in matches if x["electoral_party"] == k) for k in {x["electoral_party"] for x in matches}}).items())),
    }
    DIAGNOSTICS.write_text(json.dumps(diagnostics, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps({k: v for k, v in diagnostics.items() if k not in {"ambiguous", "unmatched_profiles", "unused_servel_elected"}}, ensure_ascii=False, indent=2))

    if len(matches) < 150:
        raise RuntimeError(f"Solo {len(matches)}/155 electos empataron con perfiles actuales")
    if ambiguous or unused:
        raise RuntimeError(
            f"La línea base requiere revisión: matched={len(matches)}, ambiguous={len(ambiguous)}, unused={len(unused)}"
        )


if __name__ == "__main__":
    main()
